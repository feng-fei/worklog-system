import openpyxl
wb = openpyxl.load_workbook('/app/data/templates/repair_template.xlsx')
ws = wb.active
print(f'Sheet: {ws.title}')
print(f'Dimensions: {ws.dimensions}')
print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
print()
print('=== Merged cells ===')
for mr in ws.merged_cells.ranges:
    print(f'  {mr}')
print()
print('=== All cells ===')
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            print(f'  R{row}C{col}: {repr(cell.value)}')
