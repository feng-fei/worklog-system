import openpyxl

wb = openpyxl.load_workbook("/tmp/repair.xlsx")
ws = wb.active
print(f"Sheets: {wb.sheetnames}, Active: {ws.title}, rows={ws.max_row}, cols={ws.max_column}")
print("--- Merged cells ---")
for mc in ws.merged_cells.ranges:
    print(f"  {mc}")
print("--- All rows ---")
for r in range(1, ws.max_row+1):
    vals = []
    for c in range(1, ws.max_column+1):
        v = ws.cell(r,c).value
        vals.append(str(v) if v is not None else "")
    print(f"  Row {r}: {vals}")

# Also check column widths, row heights, and cell styles
print("--- Column widths ---")
for i, col in enumerate(ws.column_dimensions, 1):
    if ws.column_dimensions[col].width:
        print(f"  Col {col} ({i}): width={ws.column_dimensions[col].width}")

print("--- Row heights ---")
for r in range(1, ws.max_row+1):
    if ws.row_dimensions[r].height:
        print(f"  Row {r}: height={ws.row_dimensions[r].height}")

# Check a few cell styles
print("--- Sample cell styles ---")
for r in [1, 2, 3]:
    for c in [1, 2, 3]:
        cell = ws.cell(r, c)
        if cell.value:
            print(f"  Cell({r},{c}): font={cell.font}, alignment={cell.alignment}, border={cell.border}")