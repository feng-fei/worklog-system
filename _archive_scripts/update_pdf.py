import re

with open('/app/app/routes/repair_export_routes.py', 'r', encoding='utf-8') as f:
    content = f.read()

# ===== 1. 替换 export_repair_pdf 路由中的 PDF 生成逻辑 =====
old_route_body = """        html_table = _build_html_table(ws)
        pdf_data = _build_pdf_html(html_table, record)
        pdf_bytes = _html_to_pdf(pdf_data)"""

new_route_body = """        pdf_bytes = _build_pdf_from_ws(ws, record)"""

content = content.replace(old_route_body, new_route_body)

# ===== 2. 替换 _html_to_pdf 函数为 _build_pdf_from_ws =====
old_func = """def _html_to_pdf(html_content):
    \"\"\"将 HTML 转为 PDF bytes\"\"\"
    try:
        from weasyprint import HTML
        doc = HTML(string=html_content)
        return doc.write_pdf()
    except ImportError:
        try:
            from xhtml2pdf import pisa
            output = io.BytesIO()
            pisa.CreatePDF(html_content, dest=output, encoding='utf-8')
            return output.getvalue()
        except ImportError:
            raise RuntimeError('PDF 转换库未安装，请联系管理员安装 weasyprint 或 xhtml2pdf')"""

new_func = """def _build_pdf_from_ws(ws, record):
    \"\"\"使用 fpdf2 直接从工作表构建 PDF\"\"\"
    from fpdf import FPDF
    import openpyxl

    pdf = FPDF('P', 'mm', 'A4')
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()

    font_path = '/usr/share/fonts/wqy-zenhei/wqy-zenhei.ttc'
    pdf.add_font('WQY', '', font_path)

    max_col = ws.max_column
    max_row = min(ws.max_row, 60)

    # 计算列宽（mm）
    page_w = 190
    excel_widths = []
    for col_idx in range(1, max_col + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        cd = ws.column_dimensions[col_letter]
        excel_widths.append(cd.width if cd.width else 10)
    total = sum(excel_widths)
    col_w = [page_w * w / total for w in excel_widths]

    # 计算行高（mm）
    row_h = []
    for row_idx in range(1, max_row + 1):
        rd = ws.row_dimensions[row_idx]
        row_h.append((rd.height if rd.height else 20) * 0.35)

    # 构建合并单元格查找表
    merged_master = {}
    merged_skip = set()
    for mr in ws.merged_cells.ranges:
        merged_master[(mr.min_row, mr.min_col)] = (mr.max_row, mr.max_col)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if r != mr.min_row or c != mr.min_col:
                    merged_skip.add((r, c))

    # 预计算 x 位置
    x_pos = [10]
    for w in col_w[:-1]:
        x_pos.append(x_pos[-1] + w)

    # 预计算 y 位置
    y_pos = [10]
    for h in row_h[:-1]:
        y_pos.append(y_pos[-1] + h)

    # 逐单元格绘制
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            if (row_idx, col_idx) in merged_skip:
                continue

            cell = ws.cell(row=row_idx, column=col_idx)
            val = str(cell.value) if cell.value is not None else ''

            # 单元格尺寸
            if (row_idx, col_idx) in merged_master:
                max_r, max_c = merged_master[(row_idx, col_idx)]
                cw = sum(col_w[col_idx - 1:max_c])
                ch = sum(row_h[row_idx - 1:max_r])
            else:
                cw = col_w[col_idx - 1]
                ch = row_h[row_idx - 1]

            x = x_pos[col_idx - 1]
            y = y_pos[row_idx - 1]

            # 字体样式
            font = cell.font
            fs = max(6, min(font.size if font and font.size else 9, 14))
            bold = font.bold if font else False

            # 背景色
            bg = None
            if cell.fill and cell.fill.fgColor:
                try:
                    rgb = str(cell.fill.fgColor.rgb)
                    if rgb != '00000000' and len(rgb) >= 6:
                        bg = tuple(int(rgb[i:i+2], 16) for i in range(0, 6, 2))
                except Exception:
                    pass

            # 文字颜色
            tc = (0, 0, 0)
            if font and font.color:
                try:
                    rgb = str(font.color.rgb)
                    if rgb != '00000000' and len(rgb) >= 6:
                        tc = tuple(int(rgb[i:i+2], 16) for i in range(0, 6, 2))
                except Exception:
                    pass

            # 对齐
            align = cell.alignment
            ha = 'L'
            if align and align.horizontal:
                ha = {'center': 'C', 'left': 'L', 'right': 'R'}.get(align.horizontal, 'L')

            # 绘制背景
            if bg:
                pdf.set_fill_color(*bg)
                pdf.rect(x, y, cw, ch, 'F')

            # 绘制边框
            pdf.set_draw_color(180, 180, 180)
            pdf.rect(x, y, cw, ch, 'D')

            # 绘制文字
            if val:
                pdf.set_font('WQY', '', fs)
                pdf.set_text_color(*tc)
                pdf.set_xy(x + 0.5, y + 0.3)
                pdf.multi_cell(cw - 1, fs * 0.4, val, border=0, align=ha)

    return pdf.output()"""

content = content.replace(old_func, new_func)

with open('/app/app/routes/repair_export_routes.py', 'w', encoding='utf-8') as f:
    f.write(content)

print('OK: repair_export_routes.py updated')