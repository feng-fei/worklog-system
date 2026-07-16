import io
import openpyxl
from flask import Blueprint, request, jsonify, send_file
from .. import db
from ..models import WorkRecord, SystemSetting
from ..auth import login_required

repair_export_bp = Blueprint('repair_export', __name__)

TEMPLATE_PATH = '/app/data/templates/repair_template.xlsx'


def get_company_setting(key, default=''):
    s = SystemSetting.query.filter_by(key=key).first()
    return s.value if s else default


def safe_write_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if type(cell).__name__ == 'MergedCell':
        for merged_range in ws.merged_cells.ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and
                    merged_range.min_col <= col <= merged_range.max_col):
                cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    cell.value = value


def fill_template(ws, record):
    # === 受理时间 / 维修时间 (Row 5) ===
    safe_write_cell(ws, 5, 2, record.work_date.strftime('%Y年%m月%d日') if record.work_date else '')
    time_range = ''
    if record.start_time and record.end_time:
        time_range = f"{record.start_time} - {record.end_time}"
    elif record.start_time:
        time_range = record.start_time
    safe_write_cell(ws, 5, 6, time_range)

    # === 客户信息 (Row 6-8) ===
    safe_write_cell(ws, 6, 2, record.customer_name or '')
    safe_write_cell(ws, 6, 6, record.work_subtype or '售后维修服务')
    safe_write_cell(ws, 7, 2, record.customer_phone or '')
    safe_write_cell(ws, 7, 6, record.contact_name or '')
    safe_write_cell(ws, 8, 2, record.work_address or '')
    safe_write_cell(ws, 8, 6, '')  # 保修情况

    # === 故障状况 (Row 9-13 merged) ===
    safe_write_cell(ws, 9, 1, record.fault_description or record.work_content or '')

    # === 故障分析 + 维修结果 (Row 14-22 merged) ===
    diagnosis = record.fault_diagnosis or ''
    repair_process = record.repair_process or record.remark or ''
    safe_write_cell(ws, 14, 1, f"故障分析诊断：\n{diagnosis}\n\n维修结果：\n{repair_process}")

    # === 费用清单 (Row 25-33) ===
    for r in range(25, 34):
        for c in range(2, 8):
            safe_write_cell(ws, r, c, '')

    if record.material_fee:
        safe_write_cell(ws, 25, 1, '维修材料')
        safe_write_cell(ws, 25, 2, '各类配件')
        safe_write_cell(ws, 25, 4, '批')
        safe_write_cell(ws, 25, 5, 1)
        safe_write_cell(ws, 25, 6, record.material_fee)
        safe_write_cell(ws, 25, 7, record.material_fee)

    if record.labor_fee:
        safe_write_cell(ws, 31, 1, '工时费')
        safe_write_cell(ws, 31, 2, '售后维修服务费')
        safe_write_cell(ws, 31, 4, '天/人')
        safe_write_cell(ws, 31, 5, 1)
        safe_write_cell(ws, 31, 6, record.labor_fee)
        safe_write_cell(ws, 31, 7, record.labor_fee)

    if record.transport_fee:
        safe_write_cell(ws, 32, 1, '交通费')
        safe_write_cell(ws, 32, 2, '交通费')
        safe_write_cell(ws, 32, 4, '次')
        safe_write_cell(ws, 32, 5, 1)
        safe_write_cell(ws, 32, 6, record.transport_fee)
        safe_write_cell(ws, 32, 7, record.transport_fee)

    if record.other_fee:
        safe_write_cell(ws, 33, 1, '调试费')
        safe_write_cell(ws, 33, 2, '调试费')
        safe_write_cell(ws, 33, 4, '次')
        safe_write_cell(ws, 33, 5, 1)
        safe_write_cell(ws, 33, 6, record.other_fee)
        safe_write_cell(ws, 33, 7, record.other_fee)

    total = (record.material_fee or 0) + (record.labor_fee or 0) + (record.transport_fee or 0) + (record.other_fee or 0)
    safe_write_cell(ws, 34, 6, total)

    payment_map = {'cash': '现金', 'wechat': '微信', 'alipay': '支付宝', 'transfer': '转账'}
    safe_write_cell(ws, 35, 1, f"付款方式：{payment_map.get(record.payment_status, '未指定')}")

    bank_name = get_company_setting('company_bank_name', '')
    bank_account = get_company_setting('company_bank_account', '')
    if bank_name or bank_account:
        safe_write_cell(ws, 36, 1, f"转账银行：{bank_name}  {bank_account}")

    safe_write_cell(ws, 37, 1, '')
    safe_write_cell(ws, 43, 1, '')

    staff_names = record.staff_names or record.staff_name or ''
    safe_write_cell(ws, 44, 1, f"服务人员：{staff_names}\n\n\n\n日期：\n                                     公司签章")

    company_name = get_company_setting('company_name', '珠海市瑞翼智能科技有限公司')
    company_address = get_company_setting('company_address', '')
    company_phone = get_company_setting('company_phone', '')
    company_contact = get_company_setting('company_contact', '')

    safe_write_cell(ws, 44, 5, '公司')
    safe_write_cell(ws, 44, 6, company_name)
    safe_write_cell(ws, 45, 5, '地址')
    safe_write_cell(ws, 45, 6, company_address)
    safe_write_cell(ws, 47, 5, '服务电话')
    safe_write_cell(ws, 47, 6, company_phone)
    safe_write_cell(ws, 48, 5, '联系人')
    safe_write_cell(ws, 48, 6, company_contact)
    safe_write_cell(ws, 49, 5, '服务时间')
    safe_write_cell(ws, 49, 6, '周一至周日 8:00-19:00')


# ============ 路由 ============

@repair_export_bp.route('/repair-export/<int:record_id>', methods=['GET'])
@login_required
def export_repair_order(record_id):
    record = WorkRecord.query.get(record_id)
    if not record:
        return jsonify({'error': '工单不存在'}), 404
    if record.record_type != 'repair':
        return jsonify({'error': '仅维修工单支持导出'}), 400
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb.active
        fill_template(ws, record)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"维修单_{record.order_no or record.id}_{record.customer_name}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': f'生成失败: {str(e)}'}), 500


@repair_export_bp.route('/repair-preview/<int:record_id>', methods=['GET'])
@login_required
def preview_repair_order(record_id):
    record = WorkRecord.query.get(record_id)
    if not record:
        return jsonify({'error': '工单不存在'}), 404
    if record.record_type != 'repair':
        return jsonify({'error': '仅维修工单支持预览'}), 400
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb.active
        fill_template(ws, record)
        html_table = _build_html_table(ws)
        return jsonify({'html': html_table, 'total_rows': ws.max_row, 'total_cols': ws.max_column})
    except Exception as e:
        return jsonify({'error': f'生成预览失败: {str(e)}'}), 500


@repair_export_bp.route('/repair-pdf/<int:record_id>', methods=['GET'])
@login_required
def export_repair_pdf(record_id):
    record = WorkRecord.query.get(record_id)
    if not record:
        return jsonify({'error': '工单不存在'}), 404
    if record.record_type != 'repair':
        return jsonify({'error': '仅维修工单支持导出'}), 400
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb.active
        fill_template(ws, record)
        pdf_bytes = _build_pdf_from_ws(ws, record)
        filename = f"维修单_{record.order_no or record.id}_{record.customer_name}.pdf"
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'生成PDF失败: {str(e)}'}), 500


# ============ 工具函数 ============

def _build_html_table(ws):
    """精确渲染 Excel 为 HTML 表格"""
    rows_html = []
    max_col = ws.max_column
    max_row = ws.max_row

    for row_idx in range(1, max_row + 1):
        row_height = ws.row_dimensions[row_idx].height
        row_style = f'height:{row_height * 0.75}px;' if row_height else ''
        cells_html = []
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value if cell.value is not None else ''
            val_str = str(val).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('\n', '<br>').replace('"', '&quot;')

            style_parts = []
            font = cell.font
            if font:
                if font.bold:
                    style_parts.append('font-weight:bold')
                if font.size:
                    style_parts.append(f'font-size:{font.size}pt')
                if font.color and font.color.rgb and str(font.color.rgb) != '00000000':
                    style_parts.append(f'color:#{str(font.color.rgb)[2:]}')
                if font.name:
                    style_parts.append(f'font-family:{font.name}')

            fill = cell.fill
            if fill and fill.fgColor and fill.fgColor.rgb and str(fill.fgColor.rgb) != '00000000':
                style_parts.append(f'background-color:#{str(fill.fgColor.rgb)[2:]}')

            align = cell.alignment
            if align:
                if align.horizontal:
                    h_map = {'center': 'center', 'left': 'left', 'right': 'right'}
                    style_parts.append(f'text-align:{h_map.get(align.horizontal, "left")}')
                if align.vertical:
                    v_map = {'center': 'middle', 'top': 'top', 'bottom': 'bottom'}
                    style_parts.append(f'vertical-align:{v_map.get(align.vertical, "top")}')
                if align.wrap_text:
                    style_parts.append('white-space:pre-wrap')
                if not align.wrap_text:
                    style_parts.append('white-space:nowrap')

            style_parts.append('padding:2px 6px')
            style_parts.append('border:1px solid #999')

            is_merged = False
            merge_skip = False
            for merged_range in ws.merged_cells.ranges:
                if (merged_range.min_row <= row_idx <= merged_range.max_row and
                        merged_range.min_col <= col_idx <= merged_range.max_col):
                    if row_idx == merged_range.min_row and col_idx == merged_range.min_col:
                        colspan = merged_range.max_col - merged_range.min_col + 1
                        rowspan = merged_range.max_row - merged_range.min_row + 1
                        cells_html.append(
                            f'<td colspan="{colspan}" rowspan="{rowspan}" style="{";".join(style_parts)}">{val_str}</td>'
                        )
                        is_merged = True
                    else:
                        merge_skip = True
                    break
            if merge_skip:
                continue
            if not is_merged:
                cells_html.append(f'<td style="{";".join(style_parts)}">{val_str}</td>')

        rows_html.append(f'<tr style="{row_style}">{"".join(cells_html)}</tr>')

    return f'<table style="border-collapse:collapse;width:100%;">{"".join(rows_html)}</table>'


def _build_pdf_html(html_table, record):
    """构建 PDF 专用 HTML 页面"""
    return '<!DOCTYPE html>\n<html><head><meta charset="utf-8">\n<style>\n' \
           '@page { size: A4; margin: 10mm; }\n' \
           'body { font-family: "Noto Sans CJK SC", "Noto Sans SC", "SimSun", sans-serif; font-size: 10pt; color: #000; }\n' \
           'table { border-collapse: collapse; width: 100%; }\n' \
           '.page-header { text-align: center; margin-bottom: 6px; }\n' \
           '.page-header h1 { font-size: 15pt; margin: 0 0 3px 0; }\n' \
           '.page-header p { font-size: 9pt; color: #666; margin: 0; }\n' \
           '</style></head><body>\n' \
           '<div class="page-header"><h1>维修服务单</h1>' \
           '<p>工单号: ' + str(record.order_no or record.id) + \
           ' | 日期: ' + (record.work_date.strftime('%Y-%m-%d') if record.work_date else '') + '</p></div>\n' \
           + html_table + '\n</body></html>'


def _build_pdf_from_ws(ws, record):
    """使用 fpdf2 直接从工作表构建 PDF（单页 A4）"""
    from fpdf import FPDF
    import openpyxl

    pdf = FPDF('P', 'mm', 'A4')
    pdf.set_auto_page_break(auto=False)
    pdf.add_page()

    font_path = '/usr/share/fonts/wqy-zenhei/wqy-zenhei.ttc'
    pdf.add_font('WQY', '', font_path)

    max_col = ws.max_column
    max_row = ws.max_row

    # 计算列宽（mm）
    margin = 5
    page_w = 210 - 2 * margin  # 200mm
    page_h = 297 - 2 * margin  # 287mm
    excel_widths = []
    for col_idx in range(1, max_col + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        cd = ws.column_dimensions[col_letter]
        excel_widths.append(cd.width if cd.width else 10)
    total_w = sum(excel_widths)
    col_w = [page_w * w / total_w for w in excel_widths]

    # 计算行高（mm）按比例缩放到页面高度
    excel_heights = []
    for row_idx in range(1, max_row + 1):
        rd = ws.row_dimensions[row_idx]
        excel_heights.append(rd.height if rd.height else 15)
    total_h = sum(excel_heights)
    scale = page_h / total_h
    row_h = [h * scale for h in excel_heights]

    # 构建合并单元格查找表
    merged_master = {}
    merged_skip = set()
    for mr in ws.merged_cells.ranges:
        merged_master[(mr.min_row, mr.min_col)] = (mr.max_row, mr.max_col)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if r != mr.min_row or c != mr.min_col:
                    merged_skip.add((r, c))

    # 预计算 x 位置
    x_pos = [margin]
    for w in col_w[:-1]:
        x_pos.append(x_pos[-1] + w)

    # 预计算 y 位置
    y_pos = [margin]
    for h in row_h[:-1]:
        y_pos.append(y_pos[-1] + h)

    # 逐单元格绘制
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            if (row_idx, col_idx) in merged_skip:
                continue

            cell = ws.cell(row=row_idx, column=col_idx)
            val = str(cell.value) if cell.value is not None else ''

            # 单元格尺寸
            if (row_idx, col_idx) in merged_master:
                max_r, max_c = merged_master[(row_idx, col_idx)]
                cw = sum(col_w[col_idx - 1:max_c])
                ch = sum(row_h[row_idx - 1:max_r])
            else:
                cw = col_w[col_idx - 1]
                ch = row_h[row_idx - 1]

            x = x_pos[col_idx - 1]
            y = y_pos[row_idx - 1]

            # 字体样式
            font = cell.font
            base_fs = font.size if font and font.size else 9
            fs = max(4, min(base_fs * scale * 0.9, 14))

            # 背景色
            bg = None
            if cell.fill and cell.fill.fgColor:
                try:
                    rgb = str(cell.fill.fgColor.rgb)
                    if rgb != '00000000' and len(rgb) >= 6:
                        bg = tuple(int(rgb[i:i+2], 16) for i in range(0, 6, 2))
                except Exception:
                    pass

            # 文字颜色
            tc = (0, 0, 0)
            if font and font.color:
                try:
                    rgb = str(font.color.rgb)
                    if rgb != '00000000' and len(rgb) >= 6:
                        tc = tuple(int(rgb[i:i+2], 16) for i in range(0, 6, 2))
                except Exception:
                    pass

            # 对齐
            align = cell.alignment
            ha = 'L'
            if align and align.horizontal:
                ha = {'center': 'C', 'left': 'L', 'right': 'R'}.get(align.horizontal, 'L')

            # 绘制背景
            if bg:
                pdf.set_fill_color(*bg)
                pdf.rect(x, y, cw, ch, 'F')

            # 绘制边框
            pdf.set_draw_color(180, 180, 180)
            pdf.rect(x, y, cw, ch, 'D')

            # 绘制文字
            if val:
                pdf.set_font('WQY', '', fs)
                pdf.set_text_color(*tc)
                pdf.set_xy(x + 0.5, y + 0.3)
                pdf.multi_cell(cw - 1, fs * 0.35, val, border=0, align=ha)

    return pdf.output()
