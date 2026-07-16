import openpyxl

wb = openpyxl.load_workbook("/tmp/repair.xlsx")
ws = wb.active

lines = []
lines.append(f"Sheets: {wb.sheetnames}, Active: {ws.title}, rows={ws.max_row}, cols={ws.max_column}")
lines.append("--- Merged cells ---")
for mc in ws.merged_cells.ranges:
    lines.append(f"  {mc}")
lines.append("--- All rows ---")
for r in range(1, ws.max_row+1):
    vals = []
    for c in range(1, ws.max_column+1):
        v = ws.cell(r,c).value
        vals.append(str(v) if v is not None else "")
    lines.append(f"  Row {r}: {vals}")

lines.append("--- Column widths ---")
for i, col in enumerate(ws.column_dimensions, 1):
    if ws.column_dimensions[col].width:
        lines.append(f"  Col {col}: width={ws.column_dimensions[col].width}")

lines.append("--- Row heights ---")
for r in range(1, ws.max_row+1):
    if ws.row_dimensions[r].height:
        lines.append(f"  Row {r}: height={ws.row_dimensions[r].height}")

with open("/tmp/xlsx_report.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(lines))
print("OK")