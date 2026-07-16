import openpyxl

wb = openpyxl.load_workbook("/tmp/repair.xlsx")
ws = wb.active

out = []
out.append(f"Sheets: {wb.sheetnames}, Active: {ws.title}, rows={ws.max_row}, cols={ws.max_column}")
out.append("--- Merged cells ---")
for mc in ws.merged_cells.ranges:
    out.append(f"  {mc}")
out.append("--- All rows ---")
for r in range(1, ws.max_row + 1):
    vals = [str(ws.cell(r, c).value or "") for c in range(1, ws.max_column + 1)]
    out.append(f"Row {r}: {vals}")
out.append("--- Col widths ---")
for col in ws.column_dimensions:
    if ws.column_dimensions[col].width:
        out.append(f"Col {col}: {ws.column_dimensions[col].width}")
out.append("--- Row heights ---")
for r in range(1, ws.max_row + 1):
    if ws.row_dimensions[r].height:
        out.append(f"Row {r}: {ws.row_dimensions[r].height}")

with open("/tmp/xlsx_report.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(out))

print("Report written to /tmp/xlsx_report.txt")